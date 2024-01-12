import openpyxl

fpath = r'C:\Users\LG-PC\Desktop\코딩\Python02 크롤링\05 데이터 정보 수집\02-2 참가자_data.xlsx'

# 1) 엑셀 불러오기
wb = openpyxl.load_workbook(fpath)

# 2) 엑셀 시트 선택
ws = wb['오징어게임']

# 3) 데이터 수정하기
ws['A4'] = 456
ws['B4'] = 'Gonzo'

# 4) 엑셀 저장하기
wb.save(fpath)