import openpyxl

# 1) 엑셀 만들기
wb = openpyxl.Workbook()

# 2) 엑셀 워크시트 만들기
ws = wb.create_sheet('오징어게임')

# 3) 데이터 추가하기
ws['A1'] = 'NUMBER'
ws['B1'] = 'NAME'

ws['A2'] = '001'
ws['B2'] = 'Kermit'

ws['A3'] = '002'
ws['B3'] = 'Piggy'

# 4) 엑셀 저장하기
wb.save(r'C:\Users\LG-PC\Desktop\코딩\Python02 크롤링\05 데이터 정보 수집\02 참가자_data.xlsx')
